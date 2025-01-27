from tkinter import *
from datetime import date
from tkinter import filedialog
from tkinter import messagebox
from PIL import Image, ImageTk
import os
from tkinter.ttk import Combobox
import openpyxl, xlrd
from openpyxl import Workbook
import pathlib

def login():
    global current_user
    username = entry_username.get()
    password = entry_password.get()

    if username == "" or password == "":
        messagebox.showerror("Error", "Please enter both username and password.")
    else:
        if username in user_data:
            if user_data[username] == password:
                current_user = username
                root_login.withdraw()
                root.deiconify()
            else:
                messagebox.showerror("Error", "Invalid password.")
        else:
            messagebox.showerror("Error", "Username not found.")

def register():
    username = entry_reg_username.get()
    password = entry_reg_password.get()

    if username == "" or password == "":
        messagebox.showerror("Error", "Please enter both username and password.")
    elif username in user_data:
        messagebox.showerror("Error", "Username already exists.")
    else:
        user_data[username] = password
        save_user_data()
        messagebox.showinfo("Success", "Registration successful. You can now login.")

def save_user_data():
    wb = Workbook()
    ws = wb.active
    for i, (username, password) in enumerate(user_data.items(), start=1):
        ws[f'A{i}'] = username
        ws[f'B{i}'] = password
    wb.save("user_data.xlsx")

def show_login_form_on_startup():
    root_login.deiconify()

def show_register_form():
    root_register.deiconify()

user_data = {}
if os.path.exists("user_data.xlsx"):
    wb = openpyxl.load_workbook("user_data.xlsx")
    ws = wb.active
    for row in ws.iter_rows(values_only=True):
        user_data[row[0]] = row[1]

background = "#06283D"
framebg = "#EDEDED"
framefg = "#06283D"

root=Tk()
root.title("Library System")
root.geometry("1250x700+210+100")
root.withdraw()
root.config(bg=background)
root.resizable(False, False)
file = pathlib.Path("Library_data.xlsx")
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet['A1']="Registration No."
    sheet['B1']="Name"
    sheet['C1']="Author"
    sheet['D1']="Available?"
    sheet['E1']="Date Release"
    sheet['F1']="Date of Registration"
    sheet['G1']="Category"
    sheet['H1']="Amount"

    sheet['I1']="Publisher"
    sheet['J1']="Price"
    sheet['K1']="Summary"
    sheet['L1']="Score"


    file.save("Library_data.xlsx")
#Exit
def Exit():
    root.destroy()

#Show Image
def showimage():
    global filename
    global img
    filename = filedialog.askopenfilename(initialdir=os.getcwd(), title="Select Book Image", filetypes=(("JPG File", "*.jpg"),
                                                                                                    ("PNG File", "*.png"),
                                                                                  ("All files", "*.txt")))
    img = (Image.open(filename))
    resized_image = img.resize((190, 190))
    photo2 = ImageTk.PhotoImage(resized_image)
    lbl.config(image=photo2)
    lbl.image = photo2

#Registration No.
def registration_no():
    file=openpyxl.load_workbook("Library_data.xlsx")
    sheet=file.active
    row=sheet.max_row

    max_row_value=sheet.cell(row=row, column=1).value

    try:
        Registration.set(max_row_value+1)
    except:
        Registration.set("1")

#Clear
def Clear():
    global img
    Name.set('')
    DR.set('')
    Author.set('')
    Category.set('')
    Amount.set('')
    P_Name.set('')
    Price.set('')
    Summary.set('')
    Score.set('Rating')

    registration_no()

    saveButton.config(state="normal")

    img1 = PhotoImage(file='Images/upload photo.png')
    lbl.config(image=img1)
    lbl.image = img1
    img = ""

#Save
def Save():
    R1 = Registration.get()
    N1 = Name.get()
    C1 = Category.get()
    try:
        A1=choice
    except:
        messagebox.showerror("error", "Select Available or Not!")
    D1 = DR.get()
    D2 = Date.get()
    Au1 = Author.get()
    A2 = Amount.get()
    P1 = P_Name.get()
    P2 = Price.get()
    S1 = Summary.get()
    S2 = Score.get()

    if N1 =="" or C1=="" or D1=="" or Au1=="" or A2=="" or P1=="" or P2 == "" or S1=="" or S2=="":
        messagebox.showerror("error", "Few Data is Missing!")
    else:
        file = openpyxl.load_workbook(("Library_data.xlsx"))
        sheet = file.active
        sheet.cell(column=1, row=sheet.max_row+1, value=R1)
        sheet.cell(column=2, row=sheet.max_row, value=N1)
        sheet.cell(column=3, row=sheet.max_row, value=Au1)
        sheet.cell(column=4, row=sheet.max_row, value=A1)
        sheet.cell(column=5, row=sheet.max_row, value=D1)
        sheet.cell(column=6, row=sheet.max_row, value=D2)
        sheet.cell(column=7, row=sheet.max_row, value=C1)
        sheet.cell(column=8, row=sheet.max_row, value=A1)
        sheet.cell(column=9, row=sheet.max_row, value=P1)
        sheet.cell(column=10, row=sheet.max_row, value=P2)
        sheet.cell(column=11, row=sheet.max_row, value=S1)
        sheet.cell(column=12, row=sheet.max_row, value=S2)
        file.save(r'Library_data.xlsx')

        try:
            img.save("Book Images/" +str(R1)+".jpg")
        except:
            messagebox.showerror("info", "Book Cover is not available!!!")
        messagebox.showinfo("info", "Successfully data entered!!!")
        Clear()
        registration_no()

#Delete
import openpyxl

def Delete():

    R1 = Registration.get()
    # Load the workbook
    wb = openpyxl.load_workbook("Library_data.xlsx")
    sheet = wb.active

    # Find the row to delete
    for row in sheet.iter_rows(min_row=2):  # Skip header row
        if row[0].value == R1:
            delete_row = row

    # Delete the row
    if delete_row:
        sheet.delete_rows(delete_row[0].row)
        wb.save("Library_data.xlsx")

        return True
    else:
        return False

    messagebox.showinfo("Delete Successfully")




#Search
def search():
    text = Search.get()
    Clear()
    saveButton.config(state='disable')
    file = openpyxl.load_workbook("Library_data.xlsx")
    sheet = file.active

    for row in sheet.rows:
        if row[0].value == int(text):
            name = row[0]
            # print(str(name))
            reg_no_position = str(name)[14:-1]
            reg_number = str(name)[15:-1]

    try:
        print(str(name))
    except:
        messagebox.showerror("Invalid", "Invalid registration number")

    x1 = sheet.cell(row=int(reg_number), column=1).value
    x2 = sheet.cell(row=int(reg_number), column=2).value
    x3 = sheet.cell(row=int(reg_number), column=3).value
    x4 = sheet.cell(row=int(reg_number), column=4).value
    x5 = sheet.cell(row=int(reg_number), column=5).value
    x6 = sheet.cell(row=int(reg_number), column=6).value
    x7 = sheet.cell(row=int(reg_number), column=7).value
    x8 = sheet.cell(row=int(reg_number), column=8).value
    x9 = sheet.cell(row=int(reg_number), column=9).value
    x10 = sheet.cell(row=int(reg_number), column=10).value
    x11 = sheet.cell(row=int(reg_number), column=11).value
    x12 = sheet.cell(row=int(reg_number), column=12).value

    Registration.set(x1)
    Name.set(x2)
    Author.set(x3)
    if x4 == 'NO':
        R2.select()
    else:
        R1.select()
    DR.set(x5)
    Date.set(x6)
    Category.set(x7)
    Amount.set(x8)
    P_Name.set(x9)
    Price.set(x10)
    Summary.set(x11)
    Score.set(x12)

    img = (Image.open("Book Images/"+str(x1)+".jpg"))
    resized_image = img.resize((190, 190))
    photo2 = ImageTk.PhotoImage(resized_image)
    lbl.config(image=photo2)
    lbl.image = photo2

#Update
def Update():
    R1 = Registration.get()
    N1 = Name.get()
    C1 = Category.get()
    selection()
    A1 = choice
    D1 = DR.get()
    D2 = Date.get()
    Au1 = Author.get()
    A2 = Amount.get()
    P1 = P_Name.get()
    P2 = Price.get()
    S1 = Summary.get()
    S2 = Score.get()

    file=openpyxl.load_workbook("Library_data.xlsx")
    sheet=file.active

    for row in sheet.rows:
        if row[0].value == R1:
            name=row[0]
            reg_no_position = str(name)[14:-1]
            reg_number = str(name)[15:-1]
    sheet.cell(column=1, row=int(reg_number), value=R1)
    sheet.cell(column=2, row=int(reg_number), value=N1)
    sheet.cell(column=3, row=int(reg_number), value=C1)
    sheet.cell(column=4, row=int(reg_number), value=A1)
    sheet.cell(column=5, row=int(reg_number), value=D1)
    sheet.cell(column=6, row=int(reg_number), value=D2)
    sheet.cell(column=7, row=int(reg_number), value=Au1)
    sheet.cell(column=8, row=int(reg_number), value=A2)
    sheet.cell(column=9, row=int(reg_number), value=P1)
    sheet.cell(column=10, row=int(reg_number), value=P2)
    sheet.cell(column=11, row=int(reg_number), value=S1)
    sheet.cell(column=12, row=int(reg_number), value=S2)

    file.save(r"Library_data.xlsx")

    try:
        img.save("Book Images/"+str(R1)+".jpg")
    except:
        pass
    messagebox.showinfo("Update", "Update Successfullt!!")
    Clear()
#Available
def selection():
    global choice
    value = radio.get()
    if value == 1:
        choice = "Yes"
        # print(choice)
    else:
        choice = "NO"
        # print(choice)


# Tạo cửa sổ đăng nhập
root_login = Toplevel(root)
root_login.title("Login")
root_login.geometry("300x150")

# Giao diện đăng nhập
Label(root_login, text="Username:").pack()
entry_username = Entry(root_login)
entry_username.pack()

Label(root_login, text="Password:").pack()
entry_password = Entry(root_login, show="*")
entry_password.pack()

Button(root_login, text="Login", command=login).pack()
Button(root_login, text="Register", command=show_register_form).pack()
Button(root_login, text="Exit", command=root.quit).pack()
root_login.withdraw()

# Registration form
root_register = Toplevel(root)
root_register.title("Register")
# Add your registration form elements here
# ...
# Example:
Label(root_register, text="Username:").pack()
entry_reg_username = Entry(root_register)
entry_reg_username.pack()

Label(root_register, text="Password:").pack()
entry_reg_password = Entry(root_register, show="*")
entry_reg_password.pack()

Button(root_register, text="Register", command=register).pack()
Button(root_register, text="Back to Login", command=show_login_form_on_startup).pack()
root_register.withdraw()


#--------------------------------------------------------------------------------------------------------------------------
#top frames
Label(root, text="Email: 500lambdaxl9@gmail.com", width=10, height=3, bg="#f0687c", anchor='e').pack(side=TOP, fill=X)
Label(root, text="Library System", width=10, height=2, bg="#c36464", fg="#fff", font="arial 20 bold").pack(side=TOP, fill=X)

#Search box to update
Search = StringVar()
Entry(root, textvariable=Search, width=15, bd=2, font="arial 20").place(x=820, y=70)
# imageicon3=PhotoImage(file="img/")
# Srch=Button(root, text="Search", compound=LEFT, image=imageicon3, width=123, bg="#68ddfa", font=arial 13 bold)
Srch = Button(root, text="Search", compound=LEFT, width=10, bg="#68ddfa", font="arial 13 bold", command=search)
Srch.place(x=1060, y=70)

# imageicon4=PhotoImage(file="Images/Layer 4.png")
# Update_button = Button(root, image=imageicon4, bg="#c34646", command=Update)
Update_button = Button(root, text="UPDATE", font="arial 15 bold", bg="#c34646", command=Update)
Update_button.place(x=110, y=64)
Update_button = Button(root, text="DELETE", font="arial 15 bold", bg="#c34646", command=Delete)
Update_button.place(x=230, y=64)

#Registration and Date
Label(root, text="Registration No: ", font="arial 13", fg=framebg, bg=background).place(x=30, y=150)
Label(root, text="Date: ", font="arial 13", fg=framebg, bg=background).place(x=500, y=150)

Registration = IntVar()
Date = StringVar()

reg_entry = Entry(root, textvariable=Registration, width=15, font="arial 10")
reg_entry.place(x=160, y=150)

registration_no()

today = date.today()
d1 = today.strftime("%d/%m/%Y")
date_entry = Entry(root, textvariable=Date, width=15, font="arial 10")
date_entry.place(x=550, y=150)

#Book Details
obj=LabelFrame(root, text="Book's Details", font=20, bd=2, width=900, bg=framebg, fg=framefg, height=250, relief=GROOVE)
obj.place(x=30, y=200)

Label(obj, text="Name:", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=50)
Label(obj, text="Date Release:", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=100)
Label(obj, text="Available:", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=150)


Label(obj, text="Author:", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=50)
Label(obj, text="Category:", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=100)
Label(obj, text="Amount:", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=150)


#Book Name
Name = StringVar()
name_entry = Entry(obj, textvariable=Name, width=20, font="arial 10")
name_entry.place(x=160, y=50)

#Date Release
DR = StringVar()
DR_entry = Entry(obj, textvariable=DR, width=20, font="arial 10")
DR_entry.place(x=160, y=100)

#Available
radio = IntVar()
R1 = Radiobutton(obj, text="Yes", font='arial 10 bold', variable=radio, value=1, bg=framebg, fg=framefg, command=selection)
R1.place(x=150, y=150)
R2 = Radiobutton(obj, text="Fuck NO", font='arial 10 bold', variable=radio, value=2, bg=framebg, fg=framefg, command=selection)
R2.place(x=200, y=150)

#Author Name
Author = StringVar()
author_entry = Entry(obj, textvariable=Author, width=20, font="arial 10")
author_entry.place(x=630, y=50)

#Category Name

Category = Combobox(obj, values=['Science', 'Horror', 'Literature', 'Fiction', 'History', 'Kids Book', 'Comic', 'Nonfiction', 'Romance', 'Adult'], font="Roboto 10", width=17, state='r')
Category.place(x=630, y=100)
Category.set("--Pick--")


#Amount Name
Amount = StringVar()
amount_entry = Entry(obj, textvariable=Amount, width=20, font="arial 10")
amount_entry.place(x=630, y=150)


#About
obj2 = LabelFrame(root, text="About", font=20, bd=2, width=900, bg=framebg, fg=framefg, height=220, relief=GROOVE)
obj2.place(x=30, y=470)

Label(obj2, text="Publisher:", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=50)
Label(obj2, text="Price:", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=100)

P_Name = StringVar()
p_entry = Entry(obj2, textvariable=P_Name, width=20, font="arial 10")
p_entry.place(x=160, y=50)

Price = IntVar()
price_entry = Entry(obj2, textvariable=Price, width=20, font="arial 10")
price_entry.place(x=160, y=100)

Label(obj2, text="Summary:", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=50)
Label(obj2, text="Score:", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=100)

Summary = StringVar()
summary = Entry(obj2, textvariable=Summary, width=20, font="arial 10")
summary.place(x=630, y=50)

Score = Combobox(obj2, values=['1', '2', '3', '4', '5', '6', '7', '8', '9', '10'],font="Roboto 10", width=17, state='r')
Score.place(x=630, y=100)
Score.set("--Rating--")

#Image
f = Frame(root, bd=3, bg="black", width=200, height=200, relief=GROOVE)
f.place(x=1000, y=150)

img = PhotoImage(file="Images/upload photo.png")
lbl = Label(f, bg="black", image=img)
lbl.place(x=0, y=0)

#button

Button(root, text="Upload", width=19, height=2, font="arial 12 bold", bg="lightblue", command=showimage).place(x=1000, y=370)

saveButton = Button(root, text="Save", width=19, height=2, font="arial 12 bold", bg="lightgreen", command=Save)
saveButton.place(x=1000, y=450)


Button(root, text="Reset", width=19, height=2, font="arial 12 bold", bg="lightpink", command=Clear).place(x=1000, y=530)

Button(root, text="Exit", width=19, height=2, font="arial 12 bold", bg="grey", command=Exit).place(x=1000, y=610)


Date.set(d1)

root.after(0, show_login_form_on_startup)
root.mainloop()